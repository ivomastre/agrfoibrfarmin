import openpyxl as excel
from openpyxl import Workbook
import schedule
import requests
import time
from project.tools.connect_db import *
import json
def getLista():
    # pegar os valores de uma database
    return readContacts("contacts.xlsx")

def writeContacts(fileName, cidadeNome):
    try:
        book = excel.load_workbook(fileName)
        sheet = book.active
        firstCol = sheet['A']
        print('A?'.replace("?", str(len(firstCol)+1)))
        sheet['A?'.replace("?", str(len(firstCol)+1))] = cidadeNome
        book.save("contacts.xlsx")
    except:
        book = Workbook()
        book.save("contacts.xlsx")

def readContacts(fileName):
    lst = []
    file = excel.load_workbook(fileName)
    sheet = file.active
    firstCol = sheet['A']
    for cell in range(len(firstCol)):
        contact = str(firstCol[cell].value)
        lst.append(contact)
    return lst

def job():
    lista = getLista()
    for x in lista:
        if(x=="None"):
            continue
            #gambiarra pra arrumar um bug
        print(x)
        r = requests.get("http://127.0.0.1:8000/get_cidade?cidadeNome="+x)
        print(r.text)
        dictAcao = json.loads(r.text)

        x = dictAcao
        print(x)
        create_tempo(create_connection("D:\\sqlite\db\pythonsqlite.db"), x.get("cidadeNome"), x.get("tempMin"), x.get("tempMax"), x.get("umidadeMin"), x.get("umidadeMax"))

    select_all_tempo(create_connection("D:\\sqlite\db\pythonsqlite.db"))

#schedule.every().day.at("06:00").do(job)
#job()

#while True:
    #schedule.run_pending()
    #time.sleep(1)

if __name__ == '__main__':
    job()


